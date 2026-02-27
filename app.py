"""
CRONOGRAMA UTB — app.py v2
Con módulo de finanzas, roles y exportación Excel
"""

from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg2
from psycopg2.extras import RealDictCursor
import os
from datetime import datetime
import io

# Para exportar a Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

app = Flask(__name__)

DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

SCHEMA = "cronograma"
MAX_USUARIOS = 5

# Cargos válidos del equipo (cargo = rol)
CARGOS_VALIDOS = [
    "Director de Proyecto",
    "Director de Procesos Mecanicos",
    "Director de Procesos Electronicos",
    "Diseñador de Sistemas de Control",
    "Director Financiero"
]

def cargo_a_rol(cargo):
    """Deriva el rol del sistema a partir del cargo."""
    if cargo == "Director Financiero":
        return "director_financiero"
    elif cargo == "Director de Proyecto":
        return "coordinador"
    else:
        return "miembro"


def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()
    cursor.execute(f"SET search_path TO {SCHEMA}, public")
    cursor.close()
    return conn


def init_db():
    conn = psycopg2.connect(DATABASE_URL)
    c = conn.cursor()

    c.execute(f"CREATE SCHEMA IF NOT EXISTS {SCHEMA}")
    c.execute(f"SET search_path TO {SCHEMA}, public")

    # Tabla usuarios CON ROL
    c.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id        SERIAL PRIMARY KEY,
            username  VARCHAR(100) UNIQUE NOT NULL,
            nombre    VARCHAR(200) NOT NULL,
            cargo     VARCHAR(100) NOT NULL,
            rol       VARCHAR(50) DEFAULT 'miembro',
            password  VARCHAR(255) NOT NULL,
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Tabla actividades CON detalles y observaciones
    c.execute("""
        CREATE TABLE IF NOT EXISTS actividades (
            id                 SERIAL PRIMARY KEY,
            nombre             VARCHAR(300) NOT NULL,
            descripcion        TEXT,
            detalles           TEXT,
            responsable        VARCHAR(200) NOT NULL,
            fecha_inicio       DATE NOT NULL,
            fecha_limite       DATE NOT NULL,
            prioridad          VARCHAR(20) DEFAULT 'media',
            completada         BOOLEAN DEFAULT FALSE,
            fecha_completado   DATE,
            observaciones      TEXT,
            completada_por     VARCHAR(100),
            creada_por         VARCHAR(100),
            creada_en          TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Tabla FINANZAS
    c.execute("""
        CREATE TABLE IF NOT EXISTS finanzas (
            id              SERIAL PRIMARY KEY,
            fecha_compra    DATE NOT NULL,
            concepto        VARCHAR(300) NOT NULL,
            categoria       VARCHAR(100),
            proveedor       VARCHAR(200),
            cantidad        INTEGER DEFAULT 1,
            valor_unitario  DECIMAL(12,2) NOT NULL,
            valor_total     DECIMAL(12,2) NOT NULL,
            metodo_pago     VARCHAR(50),
            responsable     VARCHAR(100),
            observaciones   TEXT,
            factura         VARCHAR(100),
            creado_por      VARCHAR(100),
            creado_en       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            modificado_por  VARCHAR(100),
            modificado_en   TIMESTAMP
        )
    """)

    c.execute("CREATE INDEX IF NOT EXISTS idx_act_fecha_limite ON actividades(fecha_limite)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_act_completada ON actividades(completada)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_usuarios_user ON usuarios(username)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_finanzas_fecha ON finanzas(fecha_compra)")

    conn.commit()
    conn.close()
    print(f"✅ BD inicializada con finanzas y roles")


try:
    init_db()
except Exception as e:
    print(f"⚠️ Error BD: {e}")


# ============================================================
# RUTAS PRINCIPALES
# ============================================================
@app.route("/")
def index():
    return render_template("index.html")


# ============================================================
# API — USUARIOS
# ============================================================
@app.route("/api/usuarios/count")
def count_usuarios():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM usuarios")
    count = c.fetchone()[0]
    conn.close()
    return jsonify({"count": count})


@app.route("/api/usuarios/listar")
def listar_usuarios():
    """Devuelve lista de usuarios con nombre y cargo para los selects de responsable."""
    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT username, nombre, cargo FROM usuarios ORDER BY nombre ASC")
    rows = c.fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/usuarios/registrar", methods=["POST"])
def registrar():
    data = request.get_json()
    nombre = data.get("nombre", "").strip()
    user   = data.get("user", "").strip().lower()
    cargo  = data.get("cargo", "").strip()
    pwd    = data.get("pass", "")

    if not all([nombre, user, cargo, pwd]):
        return jsonify({"ok": False, "error": "Faltan campos obligatorios"})

    if cargo not in CARGOS_VALIDOS:
        return jsonify({"ok": False, "error": "Cargo no válido"})

    if " " in user or not user.isalnum():
        return jsonify({"ok": False, "error": "Usuario solo puede tener letras y números sin espacios"})

    if len(pwd) < 4:
        return jsonify({"ok": False, "error": "Contraseña mínimo 4 caracteres"})

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT COUNT(*) FROM usuarios")
    if c.fetchone()[0] >= MAX_USUARIOS:
        conn.close()
        return jsonify({"ok": False, "error": f"Límite de {MAX_USUARIOS} usuarios"})

    # Verificar que el cargo no esté ya registrado (1 usuario por cargo)
    c.execute("SELECT id FROM usuarios WHERE cargo = %s", (cargo,))
    if c.fetchone():
        conn.close()
        return jsonify({"ok": False, "error": f"Ya existe un usuario con el cargo '{cargo}'"})

    c.execute("SELECT id FROM usuarios WHERE username = %s", (user,))
    if c.fetchone():
        conn.close()
        return jsonify({"ok": False, "error": "Usuario ya existe"})

    # El rol se deriva automáticamente del cargo
    rol = cargo_a_rol(cargo)

    hashed = generate_password_hash(pwd, method='pbkdf2:sha256')
    c.execute(
        "INSERT INTO usuarios (username, nombre, cargo, rol, password) VALUES (%s, %s, %s, %s, %s)",
        (user, nombre, cargo, rol, hashed)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/usuarios/login", methods=["POST"])
def login():
    data = request.get_json()
    user = data.get("user", "").strip().lower()
    pwd  = data.get("pass", "")

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM usuarios WHERE username = %s", (user,))
    row = c.fetchone()
    conn.close()

    if row and check_password_hash(row["password"], pwd):
        return jsonify({
            "ok": True,
            "usuario": {
                "id": row["id"],
                "user": row["username"],
                "nombre": row["nombre"],
                "cargo": row["cargo"],
                "rol": row["rol"]
            }
        })

    return jsonify({"ok": False, "error": "Credenciales incorrectas"})


# ============================================================
# API — ACTIVIDADES
# ============================================================
@app.route("/api/actividades", methods=["GET"])
def listar_actividades():
    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM actividades ORDER BY fecha_limite ASC")
    rows = c.fetchall()
    conn.close()

    result = []
    for r in rows:
        row_dict = dict(r)
        for campo in ['fecha_inicio', 'fecha_limite', 'fecha_completado']:
            if row_dict.get(campo):
                row_dict[campo] = str(row_dict[campo])
        result.append(row_dict)

    return jsonify(result)


@app.route("/api/actividades/<int:act_id>", methods=["GET"])
def obtener_actividad(act_id):
    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM actividades WHERE id = %s", (act_id,))
    row = c.fetchone()
    conn.close()

    if row:
        row_dict = dict(row)
        for campo in ['fecha_inicio', 'fecha_limite', 'fecha_completado']:
            if row_dict.get(campo):
                row_dict[campo] = str(row_dict[campo])
        return jsonify(row_dict)

    return jsonify({"error": "No encontrada"}), 404


@app.route("/api/actividades", methods=["POST"])
def crear_actividad():
    data = request.get_json()

    nombre = data.get("nombre", "").strip()
    responsable = data.get("responsable", "").strip()
    fecha_inicio = data.get("fecha_inicio", "")
    fecha_limite = data.get("fecha_limite", "")

    if not all([nombre, responsable, fecha_inicio, fecha_limite]):
        return jsonify({"ok": False, "error": "Faltan campos obligatorios"})

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO actividades
            (nombre, descripcion, detalles, responsable, fecha_inicio, fecha_limite, prioridad, creada_por)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        nombre,
        data.get("descripcion", ""),
        data.get("detalles", ""),
        responsable,
        fecha_inicio,
        fecha_limite,
        data.get("prioridad", "media"),
        data.get("creada_por", "")
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/actividades/<int:act_id>", methods=["PUT"])
def editar_actividad(act_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT completada FROM actividades WHERE id = %s", (act_id,))
    row = c.fetchone()

    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Actividad no encontrada"})

    if row[0]:
        conn.close()
        return jsonify({"ok": False, "error": "No se puede editar actividad completada"})

    data = request.get_json()
    c.execute("""
        UPDATE actividades
        SET nombre = %s, descripcion = %s, detalles = %s, responsable = %s,
            fecha_inicio = %s, fecha_limite = %s, prioridad = %s
        WHERE id = %s
    """, (
        data.get("nombre"),
        data.get("descripcion", ""),
        data.get("detalles", ""),
        data.get("responsable"),
        data.get("fecha_inicio"),
        data.get("fecha_limite"),
        data.get("prioridad", "media"),
        act_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/actividades/<int:act_id>/completar", methods=["PUT"])
def completar_actividad(act_id):
    data = request.get_json()
    fecha_comp = data.get("fecha_completado", "")
    observaciones = data.get("observaciones", "")
    completada_por = data.get("completada_por", "")

    if not fecha_comp:
        return jsonify({"ok": False, "error": "Falta fecha de completado"})

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT completada FROM actividades WHERE id = %s", (act_id,))
    row = c.fetchone()

    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Actividad no encontrada"})

    if row[0]:
        conn.close()
        return jsonify({"ok": False, "error": "Ya está completada"})

    c.execute("""
        UPDATE actividades
        SET completada = TRUE, fecha_completado = %s, observaciones = %s, completada_por = %s
        WHERE id = %s
    """, (fecha_comp, observaciones, completada_por, act_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/actividades/<int:act_id>", methods=["DELETE"])
def eliminar_actividad(act_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM actividades WHERE id = %s", (act_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ============================================================
# API — FINANZAS
# ============================================================
@app.route("/api/finanzas", methods=["GET"])
def listar_finanzas():
    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM finanzas ORDER BY fecha_compra DESC")
    rows = c.fetchall()
    conn.close()

    result = []
    for r in rows:
        row_dict = dict(r)
        if row_dict.get('fecha_compra'):
            row_dict['fecha_compra'] = str(row_dict['fecha_compra'])
        for campo in ['valor_unitario', 'valor_total']:
            if row_dict.get(campo):
                row_dict[campo] = float(row_dict[campo])
        result.append(row_dict)

    return jsonify(result)


@app.route("/api/finanzas", methods=["POST"])
def crear_finanza():
    data = request.get_json()

    fecha = data.get("fecha_compra", "")
    concepto = data.get("concepto", "").strip()
    valor_unitario = data.get("valor_unitario", 0)
    cantidad = data.get("cantidad", 1)

    if not all([fecha, concepto, valor_unitario]):
        return jsonify({"ok": False, "error": "Faltan campos obligatorios"})

    valor_total = float(valor_unitario) * int(cantidad)

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO finanzas
            (fecha_compra, concepto, categoria, proveedor, cantidad, valor_unitario, 
             valor_total, metodo_pago, responsable, observaciones, factura, creado_por)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        fecha,
        concepto,
        data.get("categoria", ""),
        data.get("proveedor", ""),
        cantidad,
        valor_unitario,
        valor_total,
        data.get("metodo_pago", ""),
        data.get("responsable", ""),
        data.get("observaciones", ""),
        data.get("factura", ""),
        data.get("creado_por", "")
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/finanzas/<int:fin_id>", methods=["PUT"])
def editar_finanza(fin_id):
    data = request.get_json()

    valor_unitario = data.get("valor_unitario", 0)
    cantidad = data.get("cantidad", 1)
    valor_total = float(valor_unitario) * int(cantidad)

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        UPDATE finanzas
        SET fecha_compra = %s, concepto = %s, categoria = %s, proveedor = %s,
            cantidad = %s, valor_unitario = %s, valor_total = %s,
            metodo_pago = %s, responsable = %s, observaciones = %s, factura = %s,
            modificado_por = %s, modificado_en = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (
        data.get("fecha_compra"),
        data.get("concepto"),
        data.get("categoria", ""),
        data.get("proveedor", ""),
        cantidad,
        valor_unitario,
        valor_total,
        data.get("metodo_pago", ""),
        data.get("responsable", ""),
        data.get("observaciones", ""),
        data.get("factura", ""),
        data.get("modificado_por", ""),
        fin_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/finanzas/<int:fin_id>", methods=["DELETE"])
def eliminar_finanza(fin_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM finanzas WHERE id = %s", (fin_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/finanzas/total")
def total_finanzas():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT SUM(valor_total) FROM finanzas")
    total = c.fetchone()[0] or 0
    conn.close()
    return jsonify({"total": float(total)})


# ============================================================
# EXPORTAR A EXCEL
# ============================================================
@app.route("/api/exportar/excel")
def exportar_excel():
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "openpyxl no instalado"}), 500

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_act = wb.create_sheet("Actividades")
    
    header_fill = PatternFill(start_color="003B71", end_color="003B71", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["Actividad", "Responsable (Cargo)", "Fecha Inicio", "Fecha Límite", "Fecha Completado", "Estado", "Observaciones"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_act.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM actividades ORDER BY fecha_limite ASC")
    actividades = c.fetchall()

    # Colores por estado (coinciden con leyenda de la app)
    COLOR_ESTADO = {
        'default':   'FFC107',   # Amarillo  — En ejecución
        'prematuro': '2196F3',   # Azul      — Completado prematuro
        'tiempo':    '4CAF50',   # Verde     — Completado a tiempo
        'leve':      'FF9800',   # Naranja   — Retraso leve
        'tarde':     'F44336',   # Rojo      — Retraso grave
    }

    def calcular_estado(completada, fecha_limite, fecha_completado):
        if not completada:
            return 'default', '⏳ En ejecución'
        from datetime import date
        lim  = date.fromisoformat(str(fecha_limite))
        comp = date.fromisoformat(str(fecha_completado))
        diff = (comp - lim).days
        if diff < -7:  return 'prematuro', '🔵 Prematuro'
        if diff <= 0:  return 'tiempo',    '🟢 A tiempo'
        if diff <= 7:  return 'leve',      '🟠 Retraso leve'
        return 'tarde', '🔴 Retraso grave'

    for row_num, act in enumerate(actividades, 2):
        estado_key, estado_label = calcular_estado(
            act['completada'], act['fecha_limite'], act.get('fecha_completado')
        )
        color_hex = COLOR_ESTADO[estado_key]
        estado_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        estado_font = Font(bold=True, color="FFFFFF" if estado_key in ('prematuro', 'tarde') else "000000")

        ws_act.cell(row=row_num, column=1, value=act['nombre']).border = border
        ws_act.cell(row=row_num, column=2, value=act['responsable']).border = border
        ws_act.cell(row=row_num, column=3, value=str(act['fecha_inicio'])).border = border
        ws_act.cell(row=row_num, column=4, value=str(act['fecha_limite'])).border = border
        ws_act.cell(row=row_num, column=5, value=str(act['fecha_completado']) if act.get('fecha_completado') else '').border = border

        celda_estado = ws_act.cell(row=row_num, column=6, value=estado_label)
        celda_estado.border = border
        celda_estado.fill = estado_fill
        celda_estado.font = estado_font
        celda_estado.alignment = Alignment(horizontal="center")

        ws_act.cell(row=row_num, column=7, value=act['observaciones'] or "").border = border

    ws_act.column_dimensions['A'].width = 40
    ws_act.column_dimensions['B'].width = 30
    ws_act.column_dimensions['C'].width = 15
    ws_act.column_dimensions['D'].width = 15
    ws_act.column_dimensions['E'].width = 17
    ws_act.column_dimensions['F'].width = 18
    ws_act.column_dimensions['G'].width = 40

    ws_fin = wb.create_sheet("Finanzas")
    
    fin_headers = ["Fecha", "Concepto", "Categoría", "Proveedor", "Cantidad", "V. Unitario", "V. Total", "Método Pago"]
    for col_num, header in enumerate(fin_headers, 1):
        cell = ws_fin.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    c.execute("SELECT * FROM finanzas ORDER BY fecha_compra DESC")
    finanzas = c.fetchall()
    conn.close()

    for row_num, fin in enumerate(finanzas, 2):
        ws_fin.cell(row=row_num, column=1, value=str(fin['fecha_compra'])).border = border
        ws_fin.cell(row=row_num, column=2, value=fin['concepto']).border = border
        ws_fin.cell(row=row_num, column=3, value=fin['categoria'] or "").border = border
        ws_fin.cell(row=row_num, column=4, value=fin['proveedor'] or "").border = border
        ws_fin.cell(row=row_num, column=5, value=fin['cantidad']).border = border
        ws_fin.cell(row=row_num, column=6, value=float(fin['valor_unitario'])).border = border
        ws_fin.cell(row=row_num, column=7, value=float(fin['valor_total'])).border = border
        ws_fin.cell(row=row_num, column=8, value=fin['metodo_pago'] or "").border = border

    ws_fin.column_dimensions['A'].width = 12
    ws_fin.column_dimensions['B'].width = 35
    ws_fin.column_dimensions['C'].width = 15
    ws_fin.column_dimensions['D'].width = 20
    ws_fin.column_dimensions['E'].width = 10
    ws_fin.column_dimensions['F'].width = 12
    ws_fin.column_dimensions['G'].width = 12
    ws_fin.column_dimensions['H'].width = 15

    # ── HOJA 3: GANTT ────────────────────────────────────────
    if actividades:
        from datetime import date, timedelta

        ws_g = wb.create_sheet("Gantt")

        G_BG        = "0D1B3E"
        G_HEADER    = "1A2E5C"
        G_WHITE     = "FFFFFF"
        G_SIDEBAR   = "1E3A6E"
        G_ODD       = "0F2347"
        G_EVEN      = "112A54"
        G_MONTH     = "243F72"

        # BARRA_COLORS actualizado: El verde ahora es para Electrónicos
        BARRA_COLORS = {
            "Director de Proyecto":             "2C3E50", # Azul Oscuro (Ya no es verde)
            "Director de Procesos Mecanicos":   "FF69B4", # Rosado
            "Director de Procesos Electronicos":"8DB600", # <--- AQUÍ ESTÁ TU VERDE
            "Diseñador de Sistemas de Control": "9B59B6", # Morado
            "Director Financiero":              "D4AF37", # Rojo
        }
        ESTADO_BAR = {
            "prematuro": "2196F3",
            "tiempo":    "4CAF50",
            "leve":      "FF9800",
            "tarde":     "F44336",
            "default":   "FFC107",
        }

        def g_estado(act):
            if not act["completada"]: return "default"
            lim  = date.fromisoformat(str(act["fecha_limite"]))
            comp = date.fromisoformat(str(act["fecha_completado"]))
            diff = (comp - lim).days
            if diff < -7: return "prematuro"
            if diff <= 0: return "tiempo"
            if diff <= 7: return "leve"
            return "tarde"

        g_starts = [date.fromisoformat(str(a["fecha_inicio"])) for a in actividades]
        g_ends   = [date.fromisoformat(str(a["fecha_limite"])) for a in actividades]
        p_start  = min(g_starts).replace(day=1)
        p_end    = max(g_ends)
        if p_end.month == 12:
            p_end = p_end.replace(year=p_end.year+1, month=1, day=1) - timedelta(days=1)
        else:
            p_end = p_end.replace(month=p_end.month+1, day=1) - timedelta(days=1)
        total_days = (p_end - p_start).days + 1

        COL_ACT  = 1
        COL_RESP = 2
        COL_G    = 3
        DAY_W    = 2.2
        R_TITLE  = 1
        R_MONTH  = 2
        R_DAYS   = 3
        R_DATA   = 4

        ws_g.sheet_view.showGridLines = False

        # Título
        ws_g.row_dimensions[R_TITLE].height = 36
        t_end = COL_G + total_days - 1
        ws_g.merge_cells(start_row=R_TITLE, start_column=COL_ACT,
                         end_row=R_TITLE,   end_column=t_end)
        tc = ws_g.cell(R_TITLE, COL_ACT,
                       value="CRONOGRAMA DE PROYECTO — DIAGRAMA DE GANTT · UTB")
        tc.font      = Font(name="Arial", bold=True, size=15, color=G_WHITE)
        tc.fill      = PatternFill("solid", start_color=G_BG)
        tc.alignment = Alignment(horizontal="center", vertical="center")

        ws_g.row_dimensions[R_MONTH].height = 22
        ws_g.row_dimensions[R_DAYS].height  = 16

        for label, col in [("Actividad", COL_ACT), ("Responsable", COL_RESP)]:
            ws_g.merge_cells(start_row=R_MONTH, start_column=col,
                             end_row=R_DAYS,    end_column=col)
            ch = ws_g.cell(R_MONTH, col, value=label)
            ch.fill      = PatternFill("solid", start_color=G_HEADER)
            ch.font      = Font(name="Arial", bold=True, size=10, color=G_WHITE)
            ch.alignment = Alignment(horizontal="center", vertical="center")

        # Meses y días
        cur = p_start
        col = COL_G
        while cur <= p_end:
            if cur.month == 12:
                nxt = date(cur.year+1, 1, 1)
            else:
                nxt = date(cur.year, cur.month+1, 1)
            m_end   = min(nxt - timedelta(days=1), p_end)
            days_in = (m_end - cur).days + 1

            ws_g.merge_cells(start_row=R_MONTH, start_column=col,
                             end_row=R_MONTH,   end_column=col + days_in - 1)
            mc = ws_g.cell(R_MONTH, col, value=cur.strftime("%B %Y").capitalize())
            mc.fill      = PatternFill("solid", start_color=G_MONTH)
            mc.font      = Font(name="Arial", bold=True, size=10, color=G_WHITE)
            mc.alignment = Alignment(horizontal="center", vertical="center")

            for d in range(days_in):
                dc_col  = col + d
                dc_date = cur + timedelta(days=d)
                dc = ws_g.cell(R_DAYS, dc_col, value=dc_date.day)
                dc.fill = PatternFill("solid", start_color=G_MONTH)
                dc.font = Font(name="Arial", size=7,
                               color="AAAAAA" if dc_date.weekday() < 5 else "FF6666")
                dc.alignment = Alignment(horizontal="center", vertical="center")
                ws_g.column_dimensions[get_column_letter(dc_col)].width = DAY_W

            col += days_in
            cur  = nxt

        ws_g.column_dimensions[get_column_letter(COL_ACT)].width  = 40
        ws_g.column_dimensions[get_column_letter(COL_RESP)].width  = 24

        bb = Border(bottom=Side(style="thin", color="243F72"))

        for i, act in enumerate(actividades):
            row    = R_DATA + i
            bg_row = G_ODD if i % 2 == 0 else G_EVEN
            ws_g.row_dimensions[row].height = 20

            nc = ws_g.cell(row, COL_ACT, value=act["nombre"])
            nc.fill      = PatternFill("solid", start_color=G_SIDEBAR)
            nc.font      = Font(name="Arial", size=9, color=G_WHITE)
            nc.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=1)
            nc.border    = bb

            rc = ws_g.cell(row, COL_RESP, value=act["responsable"])
            rc.fill      = PatternFill("solid", start_color=G_SIDEBAR)
            rc.font      = Font(name="Arial", size=8, color="AABBDD", italic=True)
            rc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            rc.border    = bb

            for d in range(total_days):
                gc = ws_g.cell(row, COL_G + d)
                gc.fill   = PatternFill("solid", start_color=bg_row)
                gc.border = Border(bottom=Side(style="thin", color="1A2E5C"))

            a_start = date.fromisoformat(str(act["fecha_inicio"]))
            a_end   = date.fromisoformat(str(act["fecha_limite"]))
            c_ini   = COL_G + (a_start - p_start).days
            c_fin   = COL_G + (a_end   - p_start).days

            if act["completada"]:
                bar_color = ESTADO_BAR[g_estado(act)]
            else:
                bar_color = BARRA_COLORS.get(act["responsable"], "95A5A6")

            if c_ini <= c_fin:
                ws_g.merge_cells(start_row=row, start_column=c_ini,
                                 end_row=row,   end_column=c_fin)
                bc = ws_g.cell(row, c_ini)
                bc.fill      = PatternFill("solid", start_color=bar_color)
                bc.font      = Font(name="Arial", size=7, bold=True, color=G_WHITE)
                bc.alignment = Alignment(horizontal="center", vertical="center")

        # Leyenda
        ley_row = R_DATA + len(actividades) + 2
        ws_g.row_dimensions[ley_row].height = 18
        lc = ws_g.cell(ley_row, COL_ACT, value="Actividades Completadas:")
        lc.font = Font(name="Arial", bold=True, size=9, color=G_WHITE)
        lc.fill = PatternFill("solid", start_color=G_BG)

        ley_items = [
            ("Prematuro",     ESTADO_BAR["prematuro"]),
            ("A tiempo",      ESTADO_BAR["tiempo"]),
            ("Retraso leve",  ESTADO_BAR["leve"]),
            ("Retraso grave", ESTADO_BAR["tarde"]),
        ]
        col_ley = COL_G
        for label, color in ley_items:
            ws_g.merge_cells(start_row=ley_row, start_column=col_ley,
                             end_row=ley_row,   end_column=col_ley + 9)
            box = ws_g.cell(ley_row, col_ley, value=f"  {label}  ")
            box.fill      = PatternFill("solid", start_color=color)
            box.font      = Font(name="Arial", size=8, bold=True, color=G_WHITE)
            box.alignment = Alignment(horizontal="center", vertical="center")
            col_ley += 11

        # Leyenda fila 2: Colores por cargo
        ley_row2 = ley_row + 1
        ws_g.row_dimensions[ley_row2].height = 18

        lc2 = ws_g.cell(ley_row2, COL_ACT, value="Actividades en ejecucion:")
        lc2.font = Font(name="Arial", bold=True, size=9, color=G_WHITE)
        lc2.fill = PatternFill("solid", start_color=G_BG)
        lc2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_g.cell(ley_row2, COL_RESP).fill = PatternFill("solid", start_color=G_BG)

        ley_cargos = [
            ("Dir. Proyecto",    BARRA_COLORS["Director de Proyecto"]),
            ("Dir. Proc. Mec.",  BARRA_COLORS["Director de Procesos Mecanicos"]),
            ("Dir. Proc. Elec.", BARRA_COLORS["Director de Procesos Electronicos"]),
            ("Dis. Control",     BARRA_COLORS["Diseñador de Sistemas de Control"]),
            ("Dir. Financiero",  BARRA_COLORS["Director Financiero"]),
        ]
        col_ley2 = COL_G
        for label, color in ley_cargos:
            ws_g.merge_cells(start_row=ley_row2, start_column=col_ley2,
                             end_row=ley_row2,   end_column=col_ley2 + 9)
            box = ws_g.cell(ley_row2, col_ley2, value=f"  {label}  ")
            box.fill      = PatternFill("solid", start_color=color)
            box.font      = Font(name="Arial", size=8, bold=True, color=G_WHITE)
            box.alignment = Alignment(horizontal="center", vertical="center")
            col_ley2 += 11

    

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'cronograma_utb_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ============================================================
# ADMIN — BORRAR USUARIOS
# ============================================================
@app.route("/DeleteUsers")
def delete_users_page():
    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT id, username, nombre, cargo, creado_en FROM usuarios ORDER BY creado_en ASC")
    usuarios = c.fetchall()
    conn.close()

    rows_html = ""
    for u in usuarios:
        rows_html += f"""
        <tr>
            <td>{u['nombre']}</td>
            <td>{u['username']}</td>
            <td>{u['cargo']}</td>
            <td>{str(u['creado_en'])[:10]}</td>
            <td>
                <form method="POST" action="/DeleteUsers/{u['id']}" 
                      onsubmit="return confirm('¿Eliminar a {u['nombre']}? Esta acción no se puede deshacer.')">
                    <button type="submit">🗑 Eliminar</button>
                </form>
            </td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Administrar Usuarios — UTB</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', sans-serif; background: #f4f6f9; min-height: 100vh; }}
  .header {{ background: #003B71; color: white; padding: 1.5rem 2rem; display: flex; align-items: center; gap: 1rem; }}
  .header h1 {{ font-size: 1.4rem; }}
  .header p {{ font-size: .85rem; opacity: .75; margin-top: .2rem; }}
  .container {{ max-width: 900px; margin: 2rem auto; padding: 0 1rem; }}
  .card {{ background: white; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,.08); overflow: hidden; }}
  .card-header {{ padding: 1.2rem 1.5rem; border-bottom: 1px solid #eee; display: flex; justify-content: space-between; align-items: center; }}
  .card-header h2 {{ font-size: 1.1rem; color: #003B71; }}
  .count {{ background: #e8f0fe; color: #003B71; padding: .3rem .8rem; border-radius: 20px; font-size: .85rem; font-weight: 600; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background: #f8f9fa; padding: .9rem 1.2rem; text-align: left; font-size: .82rem; color: #666; text-transform: uppercase; letter-spacing: .05em; border-bottom: 2px solid #eee; }}
  td {{ padding: .9rem 1.2rem; border-bottom: 1px solid #f0f0f0; font-size: .9rem; color: #333; vertical-align: middle; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: #fafbff; }}
  button {{ background: #c62828; color: white; border: none; padding: .45rem 1rem; border-radius: 6px; cursor: pointer; font-size: .85rem; font-weight: 600; transition: background .2s; }}
  button:hover {{ background: #b71c1c; }}
  .empty {{ text-align: center; padding: 3rem; color: #999; font-size: 1rem; }}
  .back {{ display: inline-block; margin-bottom: 1.2rem; color: #003B71; text-decoration: none; font-size: .9rem; font-weight: 600; }}
  .back:hover {{ text-decoration: underline; }}
  .alert {{ background: #e8f5e9; border: 1px solid #a5d6a7; color: #2e7d32; padding: .9rem 1.2rem; border-radius: 8px; margin-bottom: 1rem; font-weight: 600; }}
  .cargo-tag {{ background: #e8f0fe; color: #003B71; padding: .2rem .6rem; border-radius: 12px; font-size: .8rem; font-weight: 600; }}
</style>
</head>
<body>
<div class="header">
  <div>🎓</div>
  <div>
    <h1>Cronograma UTB — Administrar Usuarios</h1>
    <p>Universidad Tecnológica de Bolívar</p>
  </div>
</div>
<div class="container">
  <a class="back" href="/">← Volver a la aplicación</a>
  <div class="card">
    <div class="card-header">
      <h2>👥 Usuarios registrados</h2>
      <span class="count">{len(usuarios)} / 5</span>
    </div>
    {'<table><thead><tr><th>Nombre</th><th>Usuario</th><th>Cargo / Rol</th><th>Registrado</th><th>Acción</th></tr></thead><tbody>' + rows_html + '</tbody></table>' if usuarios else '<div class="empty">No hay usuarios registrados.</div>'}
  </div>
</div>
</body>
</html>"""


@app.route("/DeleteUsers/<int:user_id>", methods=["POST"])
def delete_user(user_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM usuarios WHERE id = %s", (user_id,))
    conn.commit()
    conn.close()
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta http-equiv="refresh" content="1;url=/DeleteUsers">
<style>
  body {{ font-family: 'Segoe UI', sans-serif; display: flex; justify-content: center; align-items: center; min-height: 100vh; background: #f4f6f9; }}
  .msg {{ background: white; padding: 2rem 3rem; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,.1); text-align: center; }}
  .msg h2 {{ color: #2e7d32; margin-bottom: .5rem; }}
  .msg p {{ color: #666; font-size: .9rem; }}
</style>
</head>
<body>
<div class="msg">
  <h2>✅ Usuario eliminado</h2>
  <p>Redirigiendo...</p>
</div>
</body>
</html>"""


# ============================================================
# INICIO
# ============================================================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)






